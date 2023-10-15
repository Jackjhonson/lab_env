import openpyxl
import matplotlib.pyplot as plt

def drawp():
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/svr_model/all_result.xlsx')
    sh = wb['训练集预测结果']
    n_rows = sh.max_row
    true_x = []
    pred_x = []
    time_y = []
    for i in range(2,n_rows+1):
        if sh.cell(i,1).value != None:
            true_x.append(sh.cell(i,3).value)
            pred_x.append(sh.cell(i,5).value)
            time_y.append(str(sh.cell(i,2).value))
        else:
            break
    
    plt.rcParams['font.sans-serif'] = ['SimHei'] # 显示中文
    plt.rcParams['axes.unicode_minus'] = False
    plt.plot(time_y,true_x,c='r')
    plt.plot(time_y,pred_x,c='y')
    plt.title('SVR模型预测值与真实值曲线（训练集，未去除UPZ和DWZ）')
    plt.xlabel('时间')
    plt.ylabel('水位')
    #保存图片
    plt.savefig('/home/zcx/python/project/code/pic_info/1.png')
    plt.show()

if __name__ == "__main__":
    drawp()