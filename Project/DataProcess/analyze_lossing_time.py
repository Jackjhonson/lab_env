# 找每张表的缺失时间，和有时间但无数据的缺失值
import json
import openpyxl

def find_time(start_time, end_time):
    s_year, s_month, s_day, s_hour, s_minute = ana_time(start_time)
    e_year, e_month, e_day, e_hour, e_minute = ana_time(end_time)
    month31 = [1,3,5,7,8,10,12]
    month30 = [4,6,9,11]
    
    all_time = []
    while not (s_year==e_year and s_month==e_month and s_day==e_day and s_hour==e_hour and s_minute==e_minute):
        merge_time_temp = merge_time(s_year, s_month, s_day, s_hour, s_minute)
        all_time.append(merge_time_temp)
        s_hour += 1
        # day进位
        if s_hour == 24:
            s_day += 1
            s_hour = 0
        # month进位
        if s_month in month30 and s_day == 31:
            s_day = 1
            s_month += 1
        elif s_month in month31 and s_day == 32:
            s_day = 1
            s_month += 1
        elif ((s_year%4==0 and s_year%100!=0) or (s_year%400==0)) and s_month == 2 and s_day == 30:
            s_day = 1
            s_month += 1
        elif (not ((s_year%4==0 and s_year%100!=0) or (s_year%400==0))) and s_month == 2 and s_day == 29:
            s_day = 1
            s_month += 1
        # year进位
        if s_month == 13:
            s_year += 1
            s_month = 1
    merge_time_temp = merge_time(e_year, e_month, e_day, e_hour, e_minute)
    all_time.append(merge_time_temp)
    return all_time

def ana_time(time):
    dt = time.split(" ")
    d = dt[0].split("-")
    t = dt[1].split(":")
    year = int(d[0])
    month = int(d[1])
    day = int(d[2])
    hour = int(t[0])
    minute = int(t[1])
    return year, month, day, hour, minute

def merge_time(year, month, day, hour, minute):
    year = str(year)
    month = str(month)
    day = str(day)
    hour = str(hour)
    minute = str(minute)
    if len(month) == 1:
        month = "0"+month
    if len(day) == 1:
        day = "0"+day
    if len(hour) == 1:
        hour = "0"+hour
    if len(minute) == 1:
        minute = "0"+minute
    return year+"-"+month+"-"+day+" "+hour+":"+minute

def ana_sheet11():
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/苏州（枫桥）.xlsx')
    all_time = find_time("2012-01-01 00:00","2022-09-28 15:00")
    sh = wb['水位']
    Z_blank_time = []
    Q_blank_time = []
    nrows = sh.max_row
    for i in range(3,nrows+1):
        dt = sh.cell(i,1).value
        TM_temp = dt.strftime("%Y-%m-%d %H:%M")
        i_time_temp = dt.strftime("%M")
        if i_time_temp != "00":
            continue
        Z_temp = sh.cell(i,2).value
        Q_temp = sh.cell(i,3).value
        all_time.remove(TM_temp)
        if Z_temp == None or Z_temp == '':
            Z_blank_time.append(TM_temp)
        if Q_temp == None or Q_temp == '':
            Q_blank_time.append(TM_temp)
    return all_time, Z_blank_time, Q_blank_time

def ana_sheet12():
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/苏州（枫桥）.xlsx')
    sh = wb['雨量']
    nrows = sh.max_row
    all_time_list = []
    STCD = []
    DRP_blank_time = []
    for i in range(2,nrows+1):
        STCD_temp = str(int(sh.cell(i,1).value))
        if STCD_temp not in STCD:
            STCD.append(STCD_temp)
            all_time_list.append(find_time("2012-01-01 00:00","2022-09-28 15:00"))
            DRP_blank_time.append([])
        dt = sh.cell(i,2).value
        TM_temp = dt.strftime("%Y-%m-%d %H:%M")
        i_time_temp = dt.strftime("%M")
        if i_time_temp != "00":
            continue
        all_time_list[STCD.index(STCD_temp)].remove(TM_temp)
        DRP_temp = sh.cell(i,3).value
        if DRP_temp == None or DRP_temp == '':
            DRP_blank_time[STCD.index(STCD_temp)].append(TM_temp)
    return STCD, all_time_list, DRP_blank_time

def ana_sheet13(iIDTM,iACCPW,iACCDW):
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/苏州（枫桥）.xlsx')
    sh = wb['沿江引排水量']
    nrows = sh.max_row
    all_time = find_time("2012-01-01 00:00","2022-09-28 15:00")
    ACCPW_blank_time = []
    ACCDW_blank_time = []
    for i in range(3,nrows+1):
        if sh.cell(i,iIDTM).value != None:
            dt = sh.cell(i,iIDTM).value
            TM_temp = dt.strftime("%Y-%m-%d %H:%M")
            i_time_temp = dt.strftime("%M")
            if i_time_temp != "00":
                continue
            ACCPW_temp = sh.cell(i,iACCPW).value
            ACCDW_temp = sh.cell(i,iACCDW).value
            all_time.remove(TM_temp)
            if ACCPW_temp == None or ACCPW_temp == '':
                ACCPW_blank_time.append(TM_temp)
            if ACCDW_temp == None or ACCDW_temp == '':
                ACCDW_blank_time.append(TM_temp)

    return all_time, ACCPW_blank_time, ACCDW_blank_time

def ana_sheet21():
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/平望.xlsx')
    all_time = find_time("2012-01-01 00:00","2022-09-28 16:00")
    sh = wb['水位']
    Z_blank_time = []
    Q_blank_time = []
    nrows = sh.max_row
    for i in range(3,nrows+1):
        dt = sh.cell(i,1).value
        TM_temp = dt.strftime("%Y-%m-%d %H:%M")
        i_time_temp = dt.strftime("%M")
        if i_time_temp != "00":
            continue
        Z_temp = sh.cell(i,2).value
        Q_temp = sh.cell(i,3).value
        all_time.remove(TM_temp)
        if Z_temp == None or Z_temp == '':
            Z_blank_time.append(TM_temp)
        if Q_temp == None or Q_temp == '':
            Q_blank_time.append(TM_temp)
    return all_time, Z_blank_time, Q_blank_time

def ana_sheet22(iSTCD,iTM,iDRP):
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/平望.xlsx')
    sh = wb['雨量']
    nrows = sh.max_row
    all_time_list = []
    STCD = []
    DRP_blank_time = []
    for i in range(3,nrows+1):
        if sh.cell(i,iTM).value != None:
            STCD_temp = str(int(sh.cell(i,iSTCD).value))
            if STCD_temp not in STCD:
                STCD.append(STCD_temp)
                all_time_list.append(find_time("2012-01-01 00:00","2022-09-28 16:00"))
                DRP_blank_time.append([])
            dt = sh.cell(i,iTM).value
            TM_temp = dt.strftime("%Y-%m-%d %H:%M")
            i_time_temp = dt.strftime("%M")
            if i_time_temp != "00":
                continue
            all_time_list[STCD.index(STCD_temp)].remove(TM_temp)
            DRP_temp = sh.cell(i,iDRP).value
            if DRP_temp == None or DRP_temp == '':
                DRP_blank_time[STCD.index(STCD_temp)].append(TM_temp)
        
    return STCD, all_time_list, DRP_blank_time

def ana_sheet23():
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/平望.xlsx')
    sh = wb['工程调度']
    nrows = sh.max_row
    all_time = find_time("2012-01-01 00:00","2022-09-28 16:00")
    UPZ_blank_time = []
    DWZ_blank_time = []
    TGTQ_blank_time = []
    for i in range(3,nrows+1):
        dt = sh.cell(i,1).value
        TM_temp = dt.strftime("%Y-%m-%d %H:%M")
        i_time_temp = dt.strftime("%M")
        if i_time_temp != "00":
            continue
        UPZ_temp = sh.cell(i,2).value
        DWZ_temp = sh.cell(i,3).value
        TGTQ_temp = sh.cell(i,4).value
        all_time.remove(TM_temp)
        if UPZ_temp == None or UPZ_temp == '':
            UPZ_blank_time.append(TM_temp)
        if DWZ_temp == None or DWZ_temp == '':
            DWZ_blank_time.append(TM_temp)
        if TGTQ_temp == None or TGTQ_temp == '':
            TGTQ_blank_time.append(TM_temp)
    return all_time, UPZ_blank_time, DWZ_blank_time, TGTQ_blank_time

def ana_sheet24():
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/平望.xlsx')
    all_time = find_time("2012-01-01 00:00","2022-09-28 16:00")
    sh = wb['潮位']
    nrows = sh.max_row
    TDZ_blank_time = []
    HLTDMK_blank_time = []
    for i in range(3,nrows+1):
        dt = sh.cell(i,1).value
        TM_temp = dt.strftime("%Y-%m-%d %H:%M")
        i_time_temp = dt.strftime("%M")
        if i_time_temp != "00":
            continue
        TDZ_temp = sh.cell(i,2).value
        HLTDMK_temp = sh.cell(i,3).value
        all_time.remove(TM_temp)
        if TDZ_temp == None or TDZ_temp == '':
            TDZ_blank_time.append(TM_temp)
        if HLTDMK_temp == None or HLTDMK_temp == '':
            HLTDMK_blank_time.append(TM_temp)
    return all_time, TDZ_blank_time, HLTDMK_blank_time

# 苏州
def generate_sheet1():
    wbw = openpyxl.Workbook()
    all_time1, Z_blank_time1, Q_blank_time1 = ana_sheet11()
    STCD2, all_time_list2, DRP_blank_time2 = ana_sheet12()
    all_time31, ACCPW_blank_time31, ACCDW_blank_time31 = ana_sheet13(1,2,3)
    all_time32, ACCPW_blank_time32, ACCDW_blank_time32 = ana_sheet13(5,6,7)
    all_time33, ACCPW_blank_time33, ACCDW_blank_time33 = ana_sheet13(9,10,11)
    all_time34, ACCPW_blank_time34, ACCDW_blank_time34 = ana_sheet13(13,14,15)
    all_time35, ACCPW_blank_time35, ACCDW_blank_time35 = ana_sheet13(17,18,19)

    shw1 = wbw.create_sheet('缺失时间')
    shw1.cell(1,1,'水位')
    stcdn = len(STCD2)
    for i in range(stcdn):
        shw1.cell(1,i+2,STCD2[i]+'(雨量)')
    shw1.cell(1,stcdn+2,'望虞闸(沿江引排水量)')
    shw1.cell(1,stcdn+3,'浒浦闸(沿江引排水量)')
    shw1.cell(1,stcdn+4,'白茆闸(沿江引排水量)')
    shw1.cell(1,stcdn+5,'七浦闸(沿江引排水量)')
    shw1.cell(1,stcdn+6,'浏河闸(沿江引排水量)')
    for i in range(len(all_time1)):
        shw1.cell(i+2,1,all_time1[i])
    for i in range(stcdn):
        at_temp = all_time_list2[i]
        for j in range(len(at_temp)):
            shw1.cell(j+2,i+2,at_temp[j])
    for i in range(len(all_time31)):
        shw1.cell(i+2,stcdn+2,all_time31[i])
    for i in range(len(all_time32)):
        shw1.cell(i+2,stcdn+3,all_time32[i])
    for i in range(len(all_time33)):
        shw1.cell(i+2,stcdn+4,all_time33[i])
    for i in range(len(all_time34)):
        shw1.cell(i+2,stcdn+5,all_time34[i])
    for i in range(len(all_time35)):
        shw1.cell(i+2,stcdn+6,all_time35[i])

    shw2 = wbw.create_sheet('缺失值时间')
    shw2.cell(1,1,'Z')
    shw2.cell(1,2,'Q')
    for i in range(stcdn):
        shw2.cell(1,i+3,'DRP('+STCD2[i]+')')
    shw2.cell(1,stcdn+3,'ACCPW(望虞闸)')
    shw2.cell(1,stcdn+4,'ACCDW(望虞闸)')
    shw2.cell(1,stcdn+5,'ACCPW(浒浦闸)')
    shw2.cell(1,stcdn+6,'ACCDW(浒浦闸)')
    shw2.cell(1,stcdn+7,'ACCPW(白茆闸)')
    shw2.cell(1,stcdn+8,'ACCDW(白茆闸)')
    shw2.cell(1,stcdn+9,'ACCPW(七浦闸)')
    shw2.cell(1,stcdn+10,'ACCDW(七浦闸)')
    shw2.cell(1,stcdn+11,'ACCPW(浏河闸)')
    shw2.cell(1,stcdn+12,'ACCDW(浏河闸)')
    for i in range(len(Z_blank_time1)):
        shw2.cell(i+2,1,Z_blank_time1[i])
    for i in range(len(Q_blank_time1)):
        shw2.cell(i+2,2,Q_blank_time1[i])
    for i in range(stcdn):
        bt_temp = DRP_blank_time2[i]
        for j in range(len(bt_temp)):
            shw2.cell(j+2,i+3,bt_temp[j])
    
    for i in range(len(ACCPW_blank_time31)):
        shw2.cell(i+2,stcdn+3,ACCPW_blank_time31[i])
    for i in range(len(ACCDW_blank_time31)):
        shw2.cell(i+2,stcdn+4,ACCDW_blank_time31[i])
    for i in range(len(ACCPW_blank_time32)):
        shw2.cell(i+2,stcdn+5,ACCPW_blank_time32[i])
    for i in range(len(ACCDW_blank_time32)):
        shw2.cell(i+2,stcdn+6,ACCDW_blank_time32[i])
    for i in range(len(ACCPW_blank_time33)):
        shw2.cell(i+2,stcdn+7,ACCPW_blank_time33[i])
    for i in range(len(ACCDW_blank_time33)):
        shw2.cell(i+2,stcdn+8,ACCDW_blank_time33[i])
    for i in range(len(ACCPW_blank_time34)):
        shw2.cell(i+2,stcdn+9,ACCPW_blank_time34[i])
    for i in range(len(ACCDW_blank_time34)):
        shw2.cell(i+2,stcdn+10,ACCDW_blank_time34[i])
    for i in range(len(ACCPW_blank_time35)):
        shw2.cell(i+2,stcdn+11,ACCPW_blank_time35[i])
    for i in range(len(ACCDW_blank_time35)):
        shw2.cell(i+2,stcdn+12,ACCDW_blank_time35[i])

    wbw.save('/home/zcx/python/project/data/lossing_time/suzhou_lossing_time.xlsx')

# 平望
def generate_sheet2():
    wbw = openpyxl.Workbook()
    all_time1, Z_blank_time1, Q_blank_time1 = ana_sheet21()
    STCD21, all_time_list21, DRP_blank_time21 = ana_sheet22(1,2,3)
    STCD22, all_time_list22, DRP_blank_time22 = ana_sheet22(7,8,9)
    all_time3, UPZ_blank_time3, DWZ_blank_time3, TGTQ_blank_time3 = ana_sheet23()
    all_time4, TDZ_blank_time4, HLTDMK_blank_time4 = ana_sheet24()

    shw1 = wbw.create_sheet('缺失时间')
    shw1.cell(1,1,'水位')
    stcdn1 = len(STCD21)
    stcdn2 = len(STCD22)
    for i in range(stcdn1):
        shw1.cell(1,i+2,STCD21[i]+'(雨量)')
    for i in range(stcdn2):
        shw1.cell(1,stcdn1+i+2,STCD22[i]+'(雨量)')
    shw1.cell(1,stcdn1+stcdn2+2,'太浦闸(工程调度)')
    shw1.cell(1,stcdn1+stcdn2+3,'米市渡潮位(潮位)')

    for i in range(len(all_time1)):
        shw1.cell(i+2,1,all_time1[i])
    for i in range(stcdn1):
        at_temp = all_time_list21[i]
        for j in range(len(at_temp)):
            shw1.cell(j+2,i+2,at_temp[j])
    for i in range(stcdn2):
        at_temp = all_time_list22[i]
        for j in range(len(at_temp)):
            shw1.cell(j+2,stcdn1+i+2,at_temp[j])
    for i in range(len(all_time3)):
        shw1.cell(i+2,stcdn1+stcdn2+2,all_time3[i])
    for i in range(len(all_time4)):
        shw1.cell(i+2,stcdn1+stcdn2+3,all_time4[i])

    shw2 = wbw.create_sheet('缺失值时间')
    shw2.cell(1,1,'Z')
    shw2.cell(1,2,'Q')
    for i in range(stcdn1):
        shw2.cell(1,i+3,'DRP('+STCD21[i]+')')
    for i in range(stcdn2):
        shw2.cell(1,stcdn1+i+3,'DRP('+STCD22[i]+')')
    shw2.cell(1,stcdn1+stcdn2+3,'UPZ(太浦闸)')
    shw2.cell(1,stcdn1+stcdn2+4,'DWZ(太浦闸)')
    shw2.cell(1,stcdn1+stcdn2+5,'TGTQ(太浦闸)')
    shw2.cell(1,stcdn1+stcdn2+6,'TDZ(米市渡潮位)')
    shw2.cell(1,stcdn1+stcdn2+7,'HLTDMK(米市渡潮位)')
    
    for i in range(len(Z_blank_time1)):
        shw2.cell(i+2,1,Z_blank_time1[i])
    for i in range(len(Q_blank_time1)):
        shw2.cell(i+2,2,Q_blank_time1[i])
    for i in range(stcdn1):
        bt_temp = DRP_blank_time21[i]
        for j in range(len(bt_temp)):
            shw2.cell(j+2,i+3,bt_temp[j])
    for i in range(stcdn2):
        bt_temp = DRP_blank_time22[i]
        for j in range(len(bt_temp)):
            shw2.cell(j+2,stcdn1+i+3,bt_temp[j])
    
    for i in range(len(UPZ_blank_time3)):
        shw2.cell(i+2,stcdn1+stcdn2+3,UPZ_blank_time3[i])
    for i in range(len(DWZ_blank_time3)):
        shw2.cell(i+2,stcdn1+stcdn2+4,DWZ_blank_time3[i])
    for i in range(len(TGTQ_blank_time3)):
        shw2.cell(i+2,stcdn1+stcdn2+5,TGTQ_blank_time3[i])
    for i in range(len(TDZ_blank_time4)):
        shw2.cell(i+2,stcdn1+stcdn2+6,TDZ_blank_time4[i])
    for i in range(len(HLTDMK_blank_time4)):
        shw2.cell(i+2,stcdn1+stcdn2+7,HLTDMK_blank_time4[i])

    wbw.save('/home/zcx/python/project/data/lossing_time/pingwang_lossing_time.xlsx')

if __name__ == "__main__":
    # suzhou_all_time = find_time("2012-01-01 00:00","2022-09-28 15:00")
    # fw1 = open('/home/zcx/python/project/data/lossing_time/suzhou_all_time.json','w')
    # json.dump(suzhou_all_time,fw1)
    # pingwang_all_time = find_time("2012-01-01 00:00","2022-09-28 16:00")
    # fw2 = open('/home/zcx/python/project/data/lossing_time/pingwang_all_time.json','w')
    # json.dump(pingwang_all_time,fw2)
    # generate_sheet1()
    generate_sheet2()