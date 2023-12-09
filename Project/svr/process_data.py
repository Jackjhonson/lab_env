import pandas as pd
import math
import json
import openpyxl

def get_data():
    ignore_columns = [3,35,37]
    ignore_rainfall_stations = [16,18,22,23,26,28,31,32]
    comp_last_columns = [2,33,34,36]
    df = pd.read_excel('/home/zcx/python/project/data/20221004平望校对.xlsx')
    n_rows = df.shape[0]
    n_columns = df.shape[1]
    # 初始窗口设置：240行（10天）+24行未来雨量+1行预测未来水位
    future_rainfall = []
    history_data = []
    predict_z = []
    for i in range(2,n_columns):
        if i in ignore_columns or i in ignore_rainfall_stations:
            continue
        temp = []
        for j in range(0,240):
            temp_data = df.iloc[j,i]
            if math.isnan(temp_data):
                if i in comp_last_columns:
                    temp_comp = temp[-1]
                    temp.append(temp_comp)
                else:
                    temp.append(0)
            else:
                temp.append(temp_data)
        history_data.append(temp)
        # 未来雨量
        if i >= 4 and i <= 32:
            future_temp = []
            for j in range(240,264):
                temp_data = df.iloc[j,i]
                if math.isnan(temp_data):
                    future_temp.append(0)
                else:
                    future_temp.append(temp_data)
            future_rainfall.append(future_temp)
        # 预测水位
        if i == 2:
            for j in range(263,0,-1):
                predict_temp = df.iloc[j,i]
                if not math.isnan(predict_temp):
                    predict_z.append(predict_temp)
                    break
    
    # 构建x
    x = []
    x_ttemp = []
    for i in range(240):
        x_temp = []
        for l in history_data:
            x_temp.append(l[i])
        x_ttemp.append(x_temp)
    for i in range(24):
        x_temp = []
        # 前面填充1个0，表示未来水位填充
        x_temp.append(0)
        for l in future_rainfall:
            x_temp.append(l[i])
        # 后面填充3个0，表示未来UPZ，DWZ，TDZ
        for _ in range(3):
            x_temp.append(0)
        x_ttemp.append(x_temp)
    x.append(x_ttemp)

    # 窗口向下滑动
    slide_step = 1
    while slide_step+263 < n_rows:
    # while slide_step+264 < 10+264:
        index1 = 0
        index2 = 0
        for i in range(2,n_columns):
            if i in ignore_columns or i in ignore_rainfall_stations:
                continue
            temp_data = df.iloc[239+slide_step,i]
            if math.isnan(temp_data):
                if i in comp_last_columns:
                    temp_comp = history_data[index1][-1]
                    history_data[index1].append(temp_comp)
                else:
                    history_data[index1].append(0)
            else:
                history_data[index1].append(temp_data)
            del history_data[index1][0]
            index1 += 1
            # 未来雨量
            if i >= 4 and i <= 32:
                temp_data = df.iloc[263+slide_step,i]
                if math.isnan(temp_data):
                    future_rainfall[index2].append(0)
                else:
                    future_rainfall[index2].append(temp_data)
                del future_rainfall[index2][0]
                index2 += 1
            # 预测水位
            if i == 2:
                temp_data = df.iloc[263+slide_step,i]
                if math.isnan(temp_data):
                    predict_z.append(predict_z[-1])
                else:
                    predict_z.append(temp_data)
        slide_step += 1
        # 构建x
        x_ttemp = []
        for i in range(240):
            x_temp = []
            for l in history_data:
                x_temp.append(l[i])
            x_ttemp.append(x_temp)
        for i in range(24):
            x_temp = []
            # 前面填充1个0，表示未来水位填充
            x_temp.append(0)
            for l in future_rainfall:
                x_temp.append(l[i])
            # 后面填充3个0，表示未来UPZ，DWZ，TDZ
            for _ in range(3):
                x_temp.append(0)
            x_ttemp.append(x_temp)
        x.append(x_ttemp)

    return x,predict_z

# 生成一维向量
def get_data2():
    ignore_columns = [3,37]
    ignore_rainfall_stations = [16,18,22,23,26,28,31,32]
    comp_last_columns = [2,33,34,35,36]
    df = pd.read_excel('/home/zcx/python/project/data/20221004平望校对.xlsx')
    n_rows = df.shape[0]
    n_columns = df.shape[1]
    # 初始窗口设置：240行（10天）+24行未来雨量+1行预测未来水位
    future_rainfall = []
    history_data = []
    predict_z = []
    for i in range(2,n_columns):
        if i in ignore_columns or i in ignore_rainfall_stations:
            continue
        temp = []
        for j in range(0,240):
            temp_data = df.iloc[j,i]
            if math.isnan(temp_data):
                if i in comp_last_columns:
                    temp_comp = temp[-1]
                    temp.append(temp_comp)
                else:
                    temp.append(0)
            else:
                temp.append(temp_data)
        history_data.append(temp)
        # 未来雨量
        if i >= 4 and i <= 32:
            future_temp = []
            for j in range(240,264):
                temp_data = df.iloc[j,i]
                if math.isnan(temp_data):
                    future_temp.append(0)
                else:
                    future_temp.append(temp_data)
            future_rainfall.append(future_temp)
        # 预测水位
        if i == 2:
            for j in range(263,0,-1):
                predict_temp = df.iloc[j,i]
                if not math.isnan(predict_temp):
                    predict_z.append(predict_temp)
                    break
    
    # 构建x
    x = []
    x_ttemp = []
    for l in history_data:
        x_ttemp.extend(l)
    for l in future_rainfall:
        x_ttemp.extend(l)
    x.append(x_ttemp)

    # 窗口向下滑动
    slide_step = 1
    while slide_step+263 < n_rows:
    # while slide_step+264 < 10+264:
        index1 = 0
        index2 = 0
        for i in range(2,n_columns):
            if i in ignore_columns or i in ignore_rainfall_stations:
                continue
            temp_data = df.iloc[239+slide_step,i]
            if math.isnan(temp_data):
                if i in comp_last_columns:
                    temp_comp = history_data[index1][-1]
                    history_data[index1].append(temp_comp)
                else:
                    history_data[index1].append(0)
            else:
                history_data[index1].append(temp_data)
            del history_data[index1][0]
            index1 += 1
            # 未来雨量
            if i >= 4 and i <= 32:
                temp_data = df.iloc[263+slide_step,i]
                if math.isnan(temp_data):
                    future_rainfall[index2].append(0)
                else:
                    future_rainfall[index2].append(temp_data)
                del future_rainfall[index2][0]
                index2 += 1
            # 预测水位
            if i == 2:
                temp_data = df.iloc[263+slide_step,i]
                if math.isnan(temp_data):
                    predict_z.append(predict_z[-1])
                else:
                    predict_z.append(temp_data)
        slide_step += 1
        # 构建x
        x_ttemp = []
        for l in history_data:
            x_ttemp.extend(l)
        for l in future_rainfall:
            x_ttemp.extend(l)
        x.append(x_ttemp)

    return x,predict_z

def save_excel(x,y):
    wbw = openpyxl.Workbook()
    shw = wbw.create_sheet('sheet1')
    l1 = len(y)
    l2 = len(x[0])
    # 表头
    for j in range(l2):
        shw.cell(1,j+1,value='x')
    shw.cell(1,l2+3,value='y')
    # 填数据
    for i in range(l1):
        for j in range(l2):
            shw.cell(i+2,j+1,value=x[i][j])
        shw.cell(i+2,l2+3,value=y[i])
    wbw.save('/home/zcx/python/project/data/one_d/pingwang_train.xlsx')

if __name__ == "__main__":
    # x,y = get_data2()
    # fw_x = open('/home/zcx/python/project/data/one_d/pingwang_x.json','w')
    # fw_y = open('/home/zcx/python/project/data/one_d/pingwang_y.json','w')
    # json.dump(x,fw_x)
    # json.dump(y,fw_y)
    fr_y = open('/home/zcx/python/project/data/one_d/pingwang_y.json','r')
    y = json.load(fr_y)
    y_csv = pd.DataFrame(data=y)
    y_csv.to_csv('/home/zcx/python/project/data/one_d/pingwang_train_y.csv')
    # x_csv = pd.DataFrame(data=x)
    # x_csv.to_csv('/home/zcx/python/project/data/one_d/pingwang_train_x.csv')
    # save_excel(x,y)