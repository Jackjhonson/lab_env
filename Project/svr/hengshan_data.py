import openpyxl
import numpy as np
import pandas as pd
import math

def find_time(start_time, end_time):
    s_year, s_month, s_day, s_hour, s_minute = ana_time(start_time)
    e_year, e_month, e_day, e_hour, e_minute = ana_time(end_time)
    month31 = [1,3,5,7,8,10,12]
    month30 = [4,6,9,11]
    
    all_time = []
    while not (s_year==e_year and s_month==e_month and s_day==e_day and s_hour==e_hour and s_minute==e_minute):
        merge_time_temp = merge_time(s_year, s_month, s_day, s_hour, s_minute)
        all_time.append(merge_time_temp)
        s_hour += 1
        # day进位
        if s_hour == 24:
            s_day += 1
            s_hour = 0
        # month进位
        if s_month in month30 and s_day == 31:
            s_day = 1
            s_month += 1
        elif s_month in month31 and s_day == 32:
            s_day = 1
            s_month += 1
        elif ((s_year%4==0 and s_year%100!=0) or (s_year%400==0)) and s_month == 2 and s_day == 30:
            s_day = 1
            s_month += 1
        elif (not ((s_year%4==0 and s_year%100!=0) or (s_year%400==0))) and s_month == 2 and s_day == 29:
            s_day = 1
            s_month += 1
        # year进位
        if s_month == 13:
            s_year += 1
            s_month = 1
    merge_time_temp = merge_time(e_year, e_month, e_day, e_hour, e_minute)
    all_time.append(merge_time_temp)
    return all_time

def ana_time(time):
    dt = time.split(" ")
    d = dt[0].split("-")
    t = dt[1].split(":")
    year = int(d[0])
    month = int(d[1])
    day = int(d[2])
    hour = int(t[0])
    minute = int(t[1])
    return year, month, day, hour, minute

def merge_time(year, month, day, hour, minute):
    year = str(year)
    month = str(month)
    day = str(day)
    hour = str(hour)
    minute = str(minute)
    if len(month) == 1:
        month = "0"+month
    if len(day) == 1:
        day = "0"+day
    if len(hour) == 1:
        hour = "0"+hour
    if len(minute) == 1:
        minute = "0"+minute
    return year+"-"+month+"-"+day+" "+hour+":"+minute

def ana_sheet1():
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/hengshan/横山水库.xlsx')
    sh = wb['水位']
    TM = []
    RZ = []
    W = []
    OTQ = []
    same_time_sum = 0
    nrows = sh.max_row
    for i in range(2,nrows+1):
        dt = sh.cell(i,1).value
        TM_min_temp = dt.strftime("%M")
        if TM_min_temp != '00':
            continue
        TM_temp = dt.strftime("%Y-%m-%d %H:%M")
        RZ_temp = sh.cell(i,2).value
        W_temp = sh.cell(i,3).value
        OTQ_temp = sh.cell(i,4).value
        if len(TM) > 0 and TM[-1] == TM_temp:
            same_time_sum += 1
            if W_temp == None:
                continue
            else:
                del TM[-1]
                del RZ[-1]
                del W[-1]
                del OTQ[-1]
        TM.append(TM_temp)
        RZ.append(RZ_temp)
        W.append(W_temp)
        OTQ.append(OTQ_temp)
    return TM, RZ, W, OTQ, same_time_sum

def ana_sheet2():
    wb = openpyxl.load_workbook('/home/zcx/python/project/data/hengshan/横山水库.xlsx')
    sh = wb['出库']
    TM = []
    DAMBHDZ = []
    DAMBHDZRCD = []
    Q = []
    S = []
    W = []
    NT = []
    nrows = sh.max_row
    for i in range(2,nrows+1):
        dt = sh.cell(i,2).value
        TM_min_temp = dt.strftime("%M")
        if TM_min_temp != '00':
            continue
        TM_temp = dt.strftime("%Y-%m-%d %H:%M")
        DAMBHDZ_temp = sh.cell(i,3).value
        DAMBHDZRCD_temp = sh.cell(i,4).value
        Q_temp = sh.cell(i,5).value
        S_temp = sh.cell(i,6).value
        W_temp = sh.cell(i,7).value
        NT_temp = sh.cell(i,8).value
        TM.append(TM_temp)
        DAMBHDZ.append(DAMBHDZ_temp)
        DAMBHDZRCD.append(DAMBHDZRCD_temp)
        Q.append(Q_temp)
        S.append(S_temp)
        W.append(W_temp)
        NT.append(NT_temp)
    return TM, DAMBHDZ, DAMBHDZRCD, Q, S, W, NT

def gen_timely_data():
    TM1, RZ1, W1, OTQ1, same_time_sum1 = ana_sheet1()
    TM2, DAMBHDZ2, DAMBHDZRCD2, Q2, S2, W2, NT2 = ana_sheet2()
    all_time = find_time("2012-01-01 08:00","2022-11-16 14:00")
    
    print("相同的时间："+str(same_time_sum1))

    # 创建sheet
    wbw = openpyxl.Workbook()
    shw = wbw.create_sheet('sheet1')
    shw.cell(1,1,'TM')
    shw.cell(1,2,'RZ')
    shw.cell(1,3,'W')
    shw.cell(1,4,'OTQ')
    shw.cell(1,5,'63101900-DAMBHDZ')
    shw.cell(1,6,'63101900-DAMBHDZRCD')
    shw.cell(1,7,'63101900-Q')
    shw.cell(1,8,'63101900-S')
    shw.cell(1,9,'63101900-W')
    shw.cell(1,10,'63101900-NT')
    # 写数据
    for i in range(len(all_time)):
        stime = all_time[i]
        shw.cell(i+2,1,str(stime))
        if TM1.__contains__(stime):
            i1 = TM1.index(stime)
            RZ_temp = RZ1[i1]
            W_temp = W1[i1]
            OTQ_temp = OTQ1[i1]
            if RZ_temp != None:
                shw.cell(i+2,2,str(RZ_temp))
            if W_temp != None:
                shw.cell(i+2,3,str(W_temp))
            if OTQ_temp != None:
                shw.cell(i+2,4,str(OTQ_temp))
        if TM2.__contains__(stime):
            i2 = TM2.index(stime)
            DAMBHDZ_temp = DAMBHDZ2[i2]
            DAMBHDZRCD_temp = DAMBHDZRCD2[i2]
            Q_temp = Q2[i2]
            S_temp = S2[i2]
            W_temp = W2[i2]
            NT_temp = NT2[i2]
            if DAMBHDZ_temp != None:
                shw.cell(i+2,5,str(DAMBHDZ_temp))
            if DAMBHDZRCD_temp != None:
                shw.cell(i+2,6,str(DAMBHDZRCD_temp))
            if Q_temp != None:
                shw.cell(i+2,7,str(Q_temp))
            if S_temp != None:
                shw.cell(i+2,8,str(S_temp))
            if W_temp != None:
                shw.cell(i+2,9,str(W_temp))
            if NT_temp != None:
                shw.cell(i+2,10,str(NT_temp))
    wbw.save('/home/zcx/python/project/data/hengshan/processed_data.xlsx')

def rainfall_data():
    d = pd.read_csv('/home/zcx/python/project/data/hengshan/横山水库雨量数据.csv')
    arr = np.array(d)
    dat = arr.tolist()

    # all_time1 = find_time("2012-01-08 08:00","2022-11-17 14:00")
    # all_time_Y1 = []
    # for i in all_time1:
    #     yy,mm,dd,hhh,mmm = split_time2(i)
    #     if str(hhh) == '08':
    #         all_time_Y1.append(i)
    # DRP1 = [0]*len(all_time1)
    # DYP1 = [0]*len(all_time_Y1)
    all_time2 = find_time("2013-07-29 16:00","2022-11-17 14:00")
    all_time_Y2 = []
    for i in all_time2:
        yy,mm,dd,hhh,mmm = split_time2(i)
        if str(hhh) == '08':
            all_time_Y2.append(i)
    DRP2 = [0]*len(all_time2)
    DYP2 = [0]*len(all_time_Y2)
    all_time3 = find_time("2013-07-29 16:00","2022-11-17 14:00")
    all_time_Y3 = []
    for i in all_time3:
        yy,mm,dd,hhh,mmm = split_time2(i)
        if str(hhh) == '08':
            all_time_Y3.append(i)
    DRP3 = [0]*len(all_time3)
    DYP3 = [0]*len(all_time_Y3)
    all_time4 = find_time("2013-07-29 16:00","2022-11-17 14:00")
    all_time_Y4 = []
    for i in all_time4:
        yy,mm,dd,hhh,mmm = split_time2(i)
        if str(hhh) == '08':
            all_time_Y4.append(i)
    DRP4 = [0]*len(all_time4)
    DYP4 = [0]*len(all_time_Y4)
    all_time5 = find_time("2013-07-29 16:00","2022-11-17 14:00")
    all_time_Y5 = []
    for i in all_time5:
        yy,mm,dd,hhh,mmm = split_time2(i)
        if str(hhh) == '08':
            all_time_Y5.append(i)
    DRP5 = [0]*len(all_time5)
    DYP5 = [0]*len(all_time_Y5)

    for rrow in dat:
        # if not math.isnan(rrow[0]) and not math.isnan(rrow[2]):
        #     time_temp1 = str(rrow[1])
        #     time_temp1 = split_time1(time_temp1)
        #     if all_time1.__contains__(time_temp1):
        #         index_temp1 = all_time1.index(time_temp1)
        #         DRP1[index_temp1] = float(rrow[2])
        #     yy,mm,dd,hhh,mmm = split_time2(time_temp1)
        #     if str(hhh) == '08':
        #         index_temp = all_time_Y1.index(time_temp1)
        #         if not math.isnan(rrow[5]):
        #             DYP1[index_temp] = float(rrow[5])
        #         else:
        #             index_temp_temp = all_time1.index(time_temp1)
        #             DYP1[index_temp] = sum(DRP1[index_temp_temp-23:index_temp_temp+1])
        if not math.isnan(rrow[7]) and not math.isnan(rrow[9]):
            time_temp2 = str(rrow[8])
            time_temp2 = split_time1(time_temp2)
            if all_time2.__contains__(time_temp2):
                index_temp2 = all_time2.index(time_temp2)
                DRP2[index_temp2] = float(rrow[9])
            yy,mm,dd,hhh,mmm = split_time2(time_temp2)
            if str(hhh) == '08':
                index_temp = all_time_Y2.index(time_temp2)
                if not math.isnan(rrow[12]):
                    DYP2[index_temp] = float(rrow[12])
                else:
                    index_temp_temp = all_time2.index(time_temp2)
                    DYP2[index_temp] = sum(DRP2[index_temp_temp-23:index_temp_temp+1])
        if not math.isnan(rrow[14]) and not math.isnan(rrow[16]):
            time_temp3 = str(rrow[15])
            time_temp3 = split_time1(time_temp3)
            if all_time3.__contains__(time_temp3):
                index_temp3 = all_time3.index(time_temp3)
                DRP3[index_temp3] = float(rrow[16])
            yy,mm,dd,hhh,mmm = split_time2(time_temp3)
            if str(hhh) == '08':
                index_temp = all_time_Y3.index(time_temp3)
                if not math.isnan(rrow[19]):
                    DYP3[index_temp] = float(rrow[19])
                else:
                    index_temp_temp = all_time3.index(time_temp3)
                    DYP3[index_temp] = sum(DRP3[index_temp_temp-23:index_temp_temp+1])
        if not math.isnan(rrow[21]) and not math.isnan(rrow[23]):
            time_temp4 = str(rrow[22])
            time_temp4 = split_time1(time_temp4)
            if all_time4.__contains__(time_temp4):
                index_temp4 = all_time4.index(time_temp4)
                DRP4[index_temp4] = float(rrow[23])
            yy,mm,dd,hhh,mmm = split_time2(time_temp4)
            if str(hhh) == '08':
                index_temp = all_time_Y4.index(time_temp4)
                if not math.isnan(rrow[26]):
                    DYP4[index_temp] = float(rrow[26])
                else:
                    index_temp_temp = all_time4.index(time_temp4)
                    DYP4[index_temp] = sum(DRP4[index_temp_temp-23:index_temp_temp+1])
        if not math.isnan(rrow[28]) and not math.isnan(rrow[30]):
            time_temp5 = str(rrow[29])
            time_temp5 = split_time1(time_temp5)
            if all_time5.__contains__(time_temp5):
                index_temp5 = all_time5.index(time_temp5)
                DRP5[index_temp5] = float(rrow[30])
            yy,mm,dd,hhh,mmm = split_time2(time_temp5)
            if str(hhh) == '08':
                index_temp = all_time_Y5.index(time_temp5)
                if not math.isnan(rrow[33]):
                    DYP5[index_temp] = float(rrow[33])
                else:
                    index_temp_temp = all_time5.index(time_temp5)
                    DYP5[index_temp] = sum(DRP5[index_temp_temp-23:index_temp_temp+1])
    
    # h_wbw = openpyxl.Workbook()
    # h_shw = h_wbw.create_sheet('sheet1')
    h_wbw = openpyxl.load_workbook('/home/zcx/python/project/data/hengshan/h_rainfall.xlsx')
    h_shw = h_wbw['sheet1']
    # h_shw.cell(1,1,'STCD')
    # h_shw.cell(1,2,'TM')
    # h_shw.cell(1,3,'DRP')
    # h_shw.cell(1,5,'STCD')
    # h_shw.cell(1,6,'TM')
    # h_shw.cell(1,7,'DRP')
    # h_shw.cell(1,9,'STCD')
    # h_shw.cell(1,10,'TM')
    # h_shw.cell(1,11,'DRP')
    # h_shw.cell(1,13,'STCD')
    # h_shw.cell(1,14,'TM')
    # h_shw.cell(1,15,'DRP')
    # h_shw.cell(1,17,'STCD')
    # h_shw.cell(1,18,'TM')
    # h_shw.cell(1,19,'DRP')
    # for i in range(len(all_time1)):
    #     h_shw.cell(i+2,1,'63101901')
    #     h_shw.cell(i+2,2,str(all_time1[i]))
    #     h_shw.cell(i+2,3,str(DRP1[i]))
    for i in range(len(all_time2)):
        h_shw.cell(i+2,5,'63124100')
        h_shw.cell(i+2,6,str(all_time2[i]))
        h_shw.cell(i+2,7,str(DRP2[i]))
    for i in range(len(all_time3)):
        h_shw.cell(i+2,9,'63124200')
        h_shw.cell(i+2,10,str(all_time3[i]))
        h_shw.cell(i+2,11,str(DRP3[i]))
    for i in range(len(all_time4)):
        h_shw.cell(i+2,13,'63124300')
        h_shw.cell(i+2,14,str(all_time4[i]))
        h_shw.cell(i+2,15,str(DRP4[i]))
    for i in range(len(all_time5)):
        h_shw.cell(i+2,17,'63124400')
        h_shw.cell(i+2,18,str(all_time5[i]))
        h_shw.cell(i+2,19,str(DRP5[i]))
    h_wbw.save('/home/zcx/python/project/data/hengshan/h_rainfall.xlsx')

    # d_wbw = openpyxl.Workbook()
    # d_shw = d_wbw.create_sheet('sheet1')
    d_wbw = openpyxl.load_workbook('/home/zcx/python/project/data/hengshan/d_rainfall.xlsx')
    d_shw = d_wbw['sheet1']
    # d_shw.cell(1,1,'STCD')
    # d_shw.cell(1,2,'TM')
    # d_shw.cell(1,3,'DYP')
    # d_shw.cell(1,5,'STCD')
    # d_shw.cell(1,6,'TM')
    # d_shw.cell(1,7,'DYP')
    # d_shw.cell(1,9,'STCD')
    # d_shw.cell(1,10,'TM')
    # d_shw.cell(1,11,'DYP')
    # d_shw.cell(1,13,'STCD')
    # d_shw.cell(1,14,'TM')
    # d_shw.cell(1,15,'DYP')
    # d_shw.cell(1,17,'STCD')
    # d_shw.cell(1,18,'TM')
    # d_shw.cell(1,19,'DYP')
    # for i in range(len(all_time_Y1)):
    #     d_shw.cell(i+2,1,'63101901')
    #     d_shw.cell(i+2,2,str(all_time_Y1[i]))
    #     d_shw.cell(i+2,3,str(DYP1[i]))
    for i in range(len(all_time_Y2)):
        d_shw.cell(i+2,5,'63124100')
        d_shw.cell(i+2,6,str(all_time_Y2[i]))
        d_shw.cell(i+2,7,str(DYP2[i]))
    for i in range(len(all_time_Y3)):
        d_shw.cell(i+2,9,'63124200')
        d_shw.cell(i+2,10,str(all_time_Y3[i]))
        d_shw.cell(i+2,11,str(DYP3[i]))
    for i in range(len(all_time_Y4)):
        d_shw.cell(i+2,13,'63124300')
        d_shw.cell(i+2,14,str(all_time_Y4[i]))
        d_shw.cell(i+2,15,str(DYP4[i]))
    for i in range(len(all_time_Y5)):
        d_shw.cell(i+2,17,'63124400')
        d_shw.cell(i+2,18,str(all_time_Y5[i]))
        d_shw.cell(i+2,19,str(DYP5[i]))
    d_wbw.save('/home/zcx/python/project/data/hengshan/d_rainfall.xlsx')

# 上述过程中，日频数据生成有问题
def rainfall_data_d():
    d = pd.read_csv('/home/zcx/python/project/data/hengshan/横山水库雨量数据.csv')
    arr = np.array(d)
    dat = arr.tolist()

    all_time1 = find_time("2012-01-08 08:00","2022-11-17 14:00")
    all_time_Y1 = []
    for i in all_time1:
        yy,mm,dd,hhh,mmm = split_time2(i)
        if str(hhh) == '08':
            all_time_Y1.append(i)
    DRP1 = [0]*len(all_time1)
    DYP1 = [0]*len(all_time_Y1)
    DYP_index1 = []
    all_time2 = find_time("2013-07-29 16:00","2022-11-17 14:00")
    all_time_Y2 = []
    for i in all_time2:
        yy,mm,dd,hhh,mmm = split_time2(i)
        if str(hhh) == '08':
            all_time_Y2.append(i)
    DRP2 = [0]*len(all_time2)
    DYP2 = [0]*len(all_time_Y2)
    DYP_index2 = []
    all_time3 = find_time("2013-07-29 16:00","2022-11-17 14:00")
    all_time_Y3 = []
    for i in all_time3:
        yy,mm,dd,hhh,mmm = split_time2(i)
        if str(hhh) == '08':
            all_time_Y3.append(i)
    DRP3 = [0]*len(all_time3)
    DYP3 = [0]*len(all_time_Y3)
    DYP_index3 = []
    all_time4 = find_time("2013-07-29 16:00","2022-11-17 14:00")
    all_time_Y4 = []
    for i in all_time4:
        yy,mm,dd,hhh,mmm = split_time2(i)
        if str(hhh) == '08':
            all_time_Y4.append(i)
    DRP4 = [0]*len(all_time4)
    DYP4 = [0]*len(all_time_Y4)
    DYP_index4 = []
    all_time5 = find_time("2013-07-29 16:00","2022-11-17 14:00")
    all_time_Y5 = []
    for i in all_time5:
        yy,mm,dd,hhh,mmm = split_time2(i)
        if str(hhh) == '08':
            all_time_Y5.append(i)
    DRP5 = [0]*len(all_time5)
    DYP5 = [0]*len(all_time_Y5)
    DYP_index5 = []

    for rrow in dat:
        if not math.isnan(rrow[0]) and not math.isnan(rrow[2]):
            time_temp1 = str(rrow[1])
            time_temp1 = split_time1(time_temp1)
            if all_time1.__contains__(time_temp1):
                index_temp1 = all_time1.index(time_temp1)
                DRP1[index_temp1] = float(rrow[2])
        if not math.isnan(rrow[0]):
            time_temp1 = str(rrow[1])
            time_temp1 = split_time1(time_temp1)
            yy,mm,dd,hhh,mmm = split_time2(time_temp1)
            if str(hhh) == '08':
                index_temp = all_time_Y1.index(time_temp1)
                if not math.isnan(rrow[5]):
                    DYP1[index_temp] = float(rrow[5])
                    DYP_index1.append(index_temp)
        if not math.isnan(rrow[7]) and not math.isnan(rrow[9]):
            time_temp2 = str(rrow[8])
            time_temp2 = split_time1(time_temp2)
            if all_time2.__contains__(time_temp2):
                index_temp2 = all_time2.index(time_temp2)
                DRP2[index_temp2] = float(rrow[9])
        if not math.isnan(rrow[7]):
            time_temp2 = str(rrow[8])
            time_temp2 = split_time1(time_temp2)
            yy,mm,dd,hhh,mmm = split_time2(time_temp2)
            if str(hhh) == '08':
                index_temp = all_time_Y2.index(time_temp2)
                if not math.isnan(rrow[12]):
                    DYP2[index_temp] = float(rrow[12])
                    DYP_index2.append(index_temp)
        if not math.isnan(rrow[14]) and not math.isnan(rrow[16]):
            time_temp3 = str(rrow[15])
            time_temp3 = split_time1(time_temp3)
            if all_time3.__contains__(time_temp3):
                index_temp3 = all_time3.index(time_temp3)
                DRP3[index_temp3] = float(rrow[16])
        if not math.isnan(rrow[14]):
            time_temp3 = str(rrow[15])
            time_temp3 = split_time1(time_temp3)
            yy,mm,dd,hhh,mmm = split_time2(time_temp3)
            if str(hhh) == '08':
                index_temp = all_time_Y3.index(time_temp3)
                if not math.isnan(rrow[19]):
                    DYP3[index_temp] = float(rrow[19])
                    DYP_index3.append(index_temp)
        if not math.isnan(rrow[21]) and not math.isnan(rrow[23]):
            time_temp4 = str(rrow[22])
            time_temp4 = split_time1(time_temp4)
            if all_time4.__contains__(time_temp4):
                index_temp4 = all_time4.index(time_temp4)
                DRP4[index_temp4] = float(rrow[23])
        if not math.isnan(rrow[21]):
            time_temp4 = str(rrow[22])
            time_temp4 = split_time1(time_temp4)
            yy,mm,dd,hhh,mmm = split_time2(time_temp4)
            if str(hhh) == '08':
                index_temp = all_time_Y4.index(time_temp4)
                if not math.isnan(rrow[26]):
                    DYP4[index_temp] = float(rrow[26])
                    DYP_index4.append(index_temp)
        if not math.isnan(rrow[28]) and not math.isnan(rrow[30]):
            time_temp5 = str(rrow[29])
            time_temp5 = split_time1(time_temp5)
            if all_time5.__contains__(time_temp5):
                index_temp5 = all_time5.index(time_temp5)
                DRP5[index_temp5] = float(rrow[30])
        if not math.isnan(rrow[28]):
            time_temp5 = str(rrow[29])
            time_temp5 = split_time1(time_temp5)
            yy,mm,dd,hhh,mmm = split_time2(time_temp5)
            if str(hhh) == '08':
                index_temp = all_time_Y5.index(time_temp5)
                if not math.isnan(rrow[33]):
                    DYP5[index_temp] = float(rrow[33])
                    DYP_index5.append(index_temp)

    for i in range(len(all_time_Y1)):
        if i in DYP_index1:
            continue
        Y_temp = all_time_Y1[i]
        index_temp_temp = all_time1.index(Y_temp)
        DYP1[i] = sum(DRP1[index_temp_temp-23:index_temp_temp+1])
    for i in range(len(all_time_Y2)):
        if i in DYP_index2:
            continue
        Y_temp = all_time_Y2[i]
        index_temp_temp = all_time2.index(Y_temp)
        DYP2[i] = sum(DRP2[index_temp_temp-23:index_temp_temp+1])
    for i in range(len(all_time_Y3)):
        if i in DYP_index3:
            continue
        Y_temp = all_time_Y3[i]
        index_temp_temp = all_time3.index(Y_temp)
        DYP3[i] = sum(DRP3[index_temp_temp-23:index_temp_temp+1])
    for i in range(len(all_time_Y4)):
        if i in DYP_index4:
            continue
        Y_temp = all_time_Y4[i]
        index_temp_temp = all_time4.index(Y_temp)
        DYP4[i] = sum(DRP4[index_temp_temp-23:index_temp_temp+1])
    for i in range(len(all_time_Y5)):
        if i in DYP_index5:
            continue
        Y_temp = all_time_Y5[i]
        index_temp_temp = all_time5.index(Y_temp)
        DYP5[i] = sum(DRP5[index_temp_temp-23:index_temp_temp+1])

    d_wbw = openpyxl.Workbook()
    d_shw = d_wbw.create_sheet('sheet1')
    # d_wbw = openpyxl.load_workbook('/home/zcx/python/project/data/hengshan/d_rainfall.xlsx')
    # d_shw = d_wbw['sheet1']
    d_shw.cell(1,1,'STCD')
    d_shw.cell(1,2,'TM')
    d_shw.cell(1,3,'DYP')
    d_shw.cell(1,5,'STCD')
    d_shw.cell(1,6,'TM')
    d_shw.cell(1,7,'DYP')
    d_shw.cell(1,9,'STCD')
    d_shw.cell(1,10,'TM')
    d_shw.cell(1,11,'DYP')
    d_shw.cell(1,13,'STCD')
    d_shw.cell(1,14,'TM')
    d_shw.cell(1,15,'DYP')
    d_shw.cell(1,17,'STCD')
    d_shw.cell(1,18,'TM')
    d_shw.cell(1,19,'DYP')
    for i in range(len(all_time_Y1)):
        d_shw.cell(i+2,1,'63101901')
        d_shw.cell(i+2,2,str(all_time_Y1[i]))
        d_shw.cell(i+2,3,str(DYP1[i]))
    for i in range(len(all_time_Y2)):
        d_shw.cell(i+2,5,'63124100')
        d_shw.cell(i+2,6,str(all_time_Y2[i]))
        d_shw.cell(i+2,7,str(DYP2[i]))
    for i in range(len(all_time_Y3)):
        d_shw.cell(i+2,9,'63124200')
        d_shw.cell(i+2,10,str(all_time_Y3[i]))
        d_shw.cell(i+2,11,str(DYP3[i]))
    for i in range(len(all_time_Y4)):
        d_shw.cell(i+2,13,'63124300')
        d_shw.cell(i+2,14,str(all_time_Y4[i]))
        d_shw.cell(i+2,15,str(DYP4[i]))
    for i in range(len(all_time_Y5)):
        d_shw.cell(i+2,17,'63124400')
        d_shw.cell(i+2,18,str(all_time_Y5[i]))
        d_shw.cell(i+2,19,str(DYP5[i]))
    d_wbw.save('/home/zcx/python/project/data/hengshan/d_rainfall.xlsx')

def split_time1(s):
    s1 = s.split(" ")
    ss2 = []
    ss1 = s1[0].split("/")
    if len(s1) == 1:
        ss2 = ['00','00']
    else:
        ss2 = s1[1].split(":")
    if len(ss1[1]) == 1:
        ss1[1] = '0'+ss1[1]
    if len(ss1[2]) == 1:
        ss1[2] = '0'+ss1[2]
    if len(ss2[0]) == 1:
        ss2[0] = '0'+ss2[0]
    time_temp = ss1[0]+"-"+ss1[1]+"-"+ss1[2]+" "+ss2[0]+":"+ss2[1]
    return time_temp

def split_time2(s):
    s1 = s.split(" ")
    ss1 = s1[0].split("-")
    ss2 = s1[1].split(":")
    return ss1[0], ss1[1], ss1[2], ss2[0], ss2[1]

if __name__ == "__main__":
    # gen_timely_data()
    # find_time("2012-01-01 08:00","2022-11-16 14:00")
    # rainfall_data()
    rainfall_data_d()